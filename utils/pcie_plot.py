from openpyxl import load_workbook
from openpyxl.chart import Series, Reference
from openpyxl.styles import Font, Color, PatternFill
from copy import deepcopy
import numpy as np

TEMPLATE_FILE = '../utils/template.xlsx'
FINAL_NAME = 'PCIE_ES2.xlsx'
DATA_LINE_INDEX = 0
SIGMA_LINE_INDEX = 7
VMIN_LINE_INDEX = 1
START_POINT_X = 11
START_POINT_Y = 5

def pcie_plot():
  workbook = load_workbook(filename = TEMPLATE_FILE, data_only= True)
  Template = get_template(workbook)
  Pattern_list = get_pattern_list(Template.specSheet)
  ptr = START_POINT_X
  # Data Plot
  for pattern in Pattern_list:
    new_sheet = workbook.create_sheet(pattern.name)
    sub_pattern1 = Template.dataSheet.cell(row = 3, column = ptr).value
    add_charts(new_sheet, ptr,      sub_pattern1, pattern, 1, Template)
    add_charts(new_sheet, ptr + 1,  sub_pattern1, pattern, 2, Template)
    sub_pattern2 = Template.dataSheet.cell(row = 3, column = ptr + 2).value  
    add_charts(new_sheet, ptr + 2,  sub_pattern2, pattern, 3, Template)
    add_charts(new_sheet, ptr + 3,  sub_pattern2, pattern, 4, Template)  
    ptr = ptr + 4
  

  Template.specSheet.sheet_state = 'hidden'
  workbook.save(r'../outputs/' + FINAL_NAME)

class Pattern:
  def __init__(self, name, condition, hot, cold):
    self.name = name
    self.condition = condition
    self.hot = hot
    self.cold = cold
    self.sigmaData = None

class TILOList:
  def __init__(self, LP, MP, HP, LP_list, MP_list, HP_list):
    self.LP = LP
    self.MP = MP
    self.HP = HP
    self.LP_list = LP_list
    self.MP_list = MP_list
    self.HP_list = HP_list

class Scale:
  def __init__(self):
    self.xmin = 0
    self.xmax = 1
    self.ymin = 0
    self.ymax = 1

class Template_Sheet:
  def __init__(self):
    self.dataSheet = None
    self.specSheet = None
    self.reportSheet = None
    self.dataLine = None
    self.sigmaLine = None
    self.vMinLine = None
    self.dashLine1 = None
    self.dasnLine2 = None
    self.LPmin = None
    self.MPmin = None
    self.HPmin = None
    self.chart = None
    self.TILO_List = None
    self.LPDash = None
    self.MPDash = None
    self.HPDash = None
    self.LP = 0
    self.MP = 0
    self.HP = 0
    self.reportPos = 0
    self.D1LP = 0
    self.D2LP = 0
    self.D1MP = 0
    self.D2MP = 0
    self.D3HP = 0
            
def get_pattern_list(worksheet):
  row = 2
  pattern_list = []
  while True:
    patternName = worksheet.cell(row = row, column = 1).value
    if patternName is None or len(patternName) == 0:
      break
    condition = worksheet.cell(row = row, column = 2).value
    hot = worksheet.cell(row = row, column = 3).value
    cold = worksheet.cell(row = row, column = 4).value
    pattern = Pattern(patternName, condition, hot, cold)
    pattern_list.append(pattern)
    row = row + 1
  print("Info: Got " + str(len(pattern_list)) + " Patterns")
  return pattern_list

def get_TILO(worksheet):
  HP_TARGET = 6
  MP_TARGET = 7
  LP_TARGET = 9
  HP = Reference(worksheet, min_col=HP_TARGET, min_row=START_POINT_Y, max_row=40)
  MP = Reference(worksheet, min_col=MP_TARGET, min_row=START_POINT_Y, max_row=40)
  LP = Reference(worksheet, min_col=LP_TARGET, min_row=START_POINT_Y, max_row=40)
  LP_list = []
  MP_list = []
  HP_list = []
  for i in range(START_POINT_Y, START_POINT_Y + 36):
    add_List(HP_list, worksheet, i, HP_TARGET)
    add_List(MP_list, worksheet, i, MP_TARGET)
    add_List(LP_list, worksheet, i, LP_TARGET)
  Tilo_List = TILOList(LP, MP, HP, LP_list, MP_list, HP_list)
  return Tilo_List

def add_List(list, worksheet, row, column):
  value = worksheet.cell(row = row, column = column).value
  if value is None:
    list.append(None)
  else:
    list.append(float(value))

def get_data(worksheet, column):
  data = Reference(worksheet, min_col = column, min_row = 5, max_row = 40)
  data_list = []
  for i in range(START_POINT_Y, START_POINT_Y + 36):
    add_List(data_list, worksheet, i, column)
  return [data, data_list]

def copy_style(line1, line2):
  line1.graphicalProperties = line2.graphicalProperties
  line1.marker = line2.marker
  line1.trendline = line2.trendline

def check_valid(worksheet, column):
  text = worksheet.cell(row = 41, column = column).value
  if text is None:
    return False
  if int(text) == 0:
    return False
  return True

def get_Vol(Vol, Template):
  TILO_list = Template.TILO_list
  TILO = None
  dashLine = None
  scale = Scale() 
  if Vol == 'LP':
    TILO = TILO_list.LP
    TILO_value_list = TILO_list.LP_list
    Vmin = Template.LPmin
    Vnum = Template.LP
    scale.xmin = 5
    scale.xmax = 10
    scale.ymin = 0.5
    scale.ymax = 0.8
    dashLine = Template.LPDash
    dashSpec = [Template.D1LP, Template.D2LP]
    dashPos = [5, 7]
  if Vol == 'MP':
    TILO = TILO_list.MP
    Vmin = Template.MPmin
    Vnum = Template.MP
    TILO_value_list = TILO_list.MP_list
    scale.xmin = 5
    scale.xmax = 10
    scale.ymin = 0.5
    scale.ymax = 0.8
    dashLine = Template.MPDash
    dashSpec = [Template.D1MP, Template.D2MP]
    dashPos = [5, 7]
  if Vol == 'HP':
    TILO = TILO_list.HP
    Vmin = Template.HPmin
    Vnum = Template.HP
    TILO_value_list = TILO_list.HP_list
    scale.xmin = 4
    scale.xmax = 7
    scale.ymin = 0.42
    scale.ymax = 0.9
    dashLine = Template.HPDash
    dashSpec = [Template.D3HP]
    dashPos = [9]
  if TILO is None:
    print("Error")
    exit(0)
  return [TILO, Vmin, Vnum, TILO_value_list, scale, dashLine, dashSpec, dashPos]

def get_location(pos, link):
  if link == True:
    if pos == 1:
      return [2, 1]
    if pos == 2:
      return [20, 1]
    if pos == 3:
      return [2, 12]
    if pos == 4:
      return [20, 12]
  if pos == 1:
    return 'B2'
  if pos == 2:
    return 'B20'
  if pos == 3:
    return 'M2'
  if pos == 4:
    return 'M20'
  
def get_template(workbook):
  Template = Template_Sheet()
  Template.specSheet = workbook['Spec']
  Template.dataSheet = workbook['Data']
  Template.reportSheet = workbook['Final Report']
  Template.chart = deepcopy(Template.specSheet._charts[0])
  Vx = Reference(Template.specSheet, min_col = 7, min_row = 11, max_row = 12)
  LPmin = Reference(Template.specSheet, min_col = 8, min_row = 11, max_row = 12)
  MPmin = Reference(Template.specSheet, min_col = 8, min_row = 13, max_row = 14)
  HPmin = Reference(Template.specSheet, min_col = 8, min_row = 15, max_row = 16)
  Template.LPmin = Series(LPmin, Vx, title = 'LPmin')
  Template.MPmin = Series(MPmin, Vx, title = 'MPmin')
  Template.HPmin = Series(HPmin, Vx, title = 'HPmin')
  Template.LP = Template.specSheet.cell(row = 11, column = 8).value
  Template.MP = Template.specSheet.cell(row = 13, column = 8).value
  Template.HP = Template.specSheet.cell(row = 15, column = 8).value
  Template.dataLine = Template.chart.series[DATA_LINE_INDEX]
  Template.vMinLine = Template.specSheet._charts[1].series[0]
  Template.sigmaLine = Template.specSheet._charts[2].series[0]
  Template.dashLine1 = Template.specSheet._charts[3].series[0]
  Template.dashLine2 = Template.specSheet._charts[4].series[0]
  copy_style(Template.LPmin, Template.vMinLine)
  copy_style(Template.MPmin, Template.vMinLine)
  copy_style(Template.HPmin, Template.vMinLine)
  Template.TILO_list = get_TILO(Template.dataSheet)
  dashliney = Reference(Template.specSheet, min_col = 8, min_row = 20, max_row = 21)
  LP_Dash1 = Reference(Template.specSheet, min_col = 7, min_row = 20, max_row = 21)
  LP_Dash2 = Reference(Template.specSheet, min_col = 7, min_row = 22, max_row = 23)
  MP_Dash1 = Reference(Template.specSheet, min_col = 7, min_row = 24, max_row = 25)
  MP_Dash2 = Reference(Template.specSheet, min_col = 7, min_row = 26, max_row = 27)
  HP_Dash3 = Reference(Template.specSheet, min_col = 7, min_row = 28, max_row = 29)
  Template.D1LP = Template.specSheet.cell(row = 20, column = 7).value
  Template.D2LP = Template.specSheet.cell(row = 22, column = 7).value
  Template.D1MP = Template.specSheet.cell(row = 24, column = 7).value
  Template.D2MP = Template.specSheet.cell(row = 26, column = 7).value
  Template.D3HP = Template.specSheet.cell(row = 28, column = 7).value
  LP_Dash1_series = Series(dashliney, LP_Dash1, title = 'dash-1LP')
  LP_Dash2_series = Series(dashliney, LP_Dash2, title = 'dash-2LP')
  MP_Dash1_series = Series(dashliney, MP_Dash1, title = 'dash-1MP')
  MP_Dash2_series = Series(dashliney, MP_Dash2, title = 'dash-2MP')
  HP_Dash3_series = Series(dashliney, HP_Dash3, title = 'dash-3HP')
  copy_style(LP_Dash1_series, Template.dashLine1)
  copy_style(LP_Dash2_series, Template.dashLine2)
  copy_style(MP_Dash1_series, Template.dashLine1)
  copy_style(MP_Dash2_series, Template.dashLine2)
  copy_style(HP_Dash3_series, Template.dashLine1)
  Template.LPDash = [LP_Dash1_series, LP_Dash2_series]
  Template.MPDash = [MP_Dash1_series, MP_Dash2_series]
  Template.HPDash = [HP_Dash3_series]
  return Template

def set_title(title, text):
  title.text.rich.p[0].r[0].t = text

def add_sigma(worksheet, ptr, Template, pos, x_list, y_list, dashSpec):
  x_clean_list = []
  y_clean_list = []
  for i in range(0, 36):
    if x_list[i] is None:
      continue
    if y_list[i] is None:
      continue
    x_clean_list.append(x_list[i])
    y_clean_list.append(y_list[i])
  # cal the std
  x = np.array(x_clean_list)
  y = np.array(y_clean_list)
  [a, b] = np.polyfit(x, np.log(y), 1)
  diff = []
  for i in range(0, len(x_clean_list)):
    y_est = np.exp(b) * np.exp(a * x_clean_list[i])
    diff.append(y_clean_list[i] - y_est)
  stdev = np.std(diff)
 # gen 3-sigma
  sigma_y = []
  for i in range(0, 36):
    value = Template.dataSheet.cell(row = START_POINT_Y + i, column = ptr).value
    if not (value is None):
      sigma = 3 * stdev + float(value)
      worksheet.cell(row = START_POINT_Y + i, column = pos + 24).value = sigma
      sigma_y.append(sigma)
    else :
      sigma_y.append(None)
  # gen the 3-sigma equation
  x_clean_list = []
  y_clean_list = []
  for i in range(0, 36):
    if x_list[i] is None:
      continue
    if sigma_y[i] is None:
      continue
    x_clean_list.append(x_list[i])
    y_clean_list.append(sigma_y[i])
  x = np.array(x_clean_list)
  y = np.array(y_clean_list)
  [a, b] = np.polyfit(x, np.log(y), 1)
  crossV = []
  for i in range(0, len(dashSpec)):
    vol = float(dashSpec[i])
    y_est = np.exp(b) * np.exp(a * vol)
    crossV.append(y_est)
  return crossV

def plot_sigma(worksheet, Template, pos, xvalues):
  data = get_data(worksheet, pos + 24)[0]
  series = Series(data, xvalues, title = '3-sigma')
  copy_style(series, Template.sigmaLine)
  return series

def add_link(worksheet,pos, reportPos):
  cell = worksheet.cell(row = pos[0], column = pos[1])
  cell.value = 'Report'
  cell.font = Font(bold = True)
  # cell.fill = PatternFill(start_color="69c0ff", end_color="69c0ff", fill_type = "solid")
  link = "#\'Final Report\'!C" + str(reportPos + 6) + ':J' + str(reportPos + 6)
  cell.hyperlink = link
  
def add_charts(worksheet, ptr, subName, pattern, pos, Template):
  if not check_valid(Template.dataSheet, ptr):
    return
  [data, data_list] = get_data(Template.dataSheet, ptr)
  if pos % 2 == 1:
    temp = pattern.hot
  else :
    temp = pattern.cold
  [xvalues, Vmin, Vnum, TILO_value_list, scale, dashLine, dashSpec, dashPos] = get_Vol(pattern.condition, Template)
  new_chart = deepcopy(Template.chart)
  # Data Plot
  series = Series(data, xvalues, title = subName + ' ' + str(temp))
  copy_style(series, Template.dataLine)
  new_chart.series[0] = series
  new_chart.series.append(Vmin)
  # 3 sigma Line
  crossV = add_sigma(worksheet, ptr, Template, pos, TILO_value_list, data_list, dashSpec)
  sigma_line = plot_sigma(worksheet, Template, pos, xvalues)
  new_chart.series.append(sigma_line)
  for dash in dashLine:
    new_chart.append(dash)
  # Chart
  new_chart.height = 9
  new_chart.width = 16
  new_chart.x_axis.scaling.min = scale.xmin
  new_chart.x_axis.scaling.max = scale.xmax
  new_chart.y_axis.scaling.min = scale.ymin
  new_chart.y_axis.scaling.max = scale.ymax
  title = 'Versal VC1902 ES2 : ' + pattern.name + ' ' + subName + ' @ ' + str(temp)
  set_title(new_chart.title, title)
  x_axis_title = 'TILO,' + str(temp) + ' @ ' + str(Vnum) + ' VCCINT'
  set_title(new_chart.x_axis.title, x_axis_title)
  worksheet.add_chart(new_chart, get_location(pos, False))
  add_link(worksheet, get_location(pos, True), Template.reportPos)
  # Report Page
  # margin = []
  # for cross in crossV:
  #   margin.append(Vnum - cross)
  add_report(pattern, subName, Template, temp, crossV, float(Vnum), dashPos)    
  
def add_report(pattern, subName, Template, temp, crossV, Vnum, dashPos):
  reportSheet = Template.reportSheet
  row = Template.reportPos + 6
  Template.reportPos = Template.reportPos + 1
  if temp > 0:
    reportSheet.cell(row = row, column = 2).value = pattern.name + ' ' + subName
  reportSheet.cell(row = row, column = 3).value = temp
  reportSheet.cell(row = row, column = 4).value = pattern.condition
  margin = []
  link = "#\'" + pattern.name +"\'!A1"
  reportSheet.cell(row = row, column = 2).hyperlink  = link
  for i in range(0, len(crossV)):
    margin = Vnum - crossV[i]
    reportSheet.cell(row = row, column = dashPos[i]).value = margin * 1000
    reportSheet.cell(row = row, column = dashPos[i] + 1).value = margin / Vnum
    if margin < 0:
      ft = Font(color="FF0000")
      reportSheet.cell(row = row, column = dashPos[i]).font = ft
      reportSheet.cell(row = row, column = dashPos[i] + 1).font = ft
  if temp < 0:
    reportSheet.merge_cells(start_row = row - 1, start_column = 2, end_row = row, end_column = 2)
    for i in range(3, 11):
      reportSheet.cell(row = row, column = i).fill = PatternFill(start_color="f0f5ff", end_color="f0f5ff", fill_type = "solid")
pcie_plot()
