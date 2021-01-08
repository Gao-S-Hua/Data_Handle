from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference, LineChart
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.trendline import Trendline, TrendlineLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.text import Text
#### CONSTANTS SETTINGS #############
CHART_LOCATION = "B10"

def gen_excel(test_result, spec = 100):
  wb = Workbook()
  summary = gen_summary(test_result)
  add_summarize(summary, wb)
  for result in test_result:
    ws = wb.create_sheet(result.SN + "_" + result.temperature)
    title = ["SN", "Temperature", "Fmax", "Spec"]
    ws.append(title)
    for data in result.data:
      ws.append([result.SN, result.temperature, data, spec])
    add_fmax_plot(ws, len(result.data))
  page_one=wb.get_sheet_by_name('Sheet')
  wb.remove_sheet(page_one)
  wb.save(r"./outputs/Result.xlsx")
  print("| Result.xlsx has been generated")
  return summary

def gen_summary(test_result):
  pass_count = 0
  fail_count = 0
  result_count = 0
  length_array = []
  for result in test_result:
    result_count = result_count + 1
    length_array.append(len(result.data))
    for data in result.data:
      if data > 100 :
        pass_count = pass_count + 1
      else:
        fail_count = fail_count + 1

  my_summary = Summary(pass_count, fail_count, result_count,length_array)
  return my_summary


def add_summarize(summary, wb):
  ws = wb.create_sheet("Summary")
  title = ["Pass", "Fail"]
  ws.append(title)
  ws.append([summary.pass_count, summary.fail_count])
  add_counter(ws)
  return wb

def add_fmax_plot(ws, length):
  this_chart = LineChart()
  this_chart.title = "PCIE Fmax"
  this_chart.style = 13
  this_chart.y_axis.title = 'Frequency (MHz)'
  this_chart.x_axis.title = 'Test count'
  data = Reference(ws, min_col=3, min_row=1, max_col=4, max_row=8)
  this_chart.add_data(data, titles_from_data=True)

  data_line = this_chart.series[0]
  set_data_line(data_line)

  spec_line = this_chart.series[1]
  set_spec_line(spec_line)
  ws.add_chart(this_chart, CHART_LOCATION)

def add_counter(ws):
  chart1 = BarChart()
  chart1.type = "col"
  chart1.style = 10
  chart1.title = "PCIE Test Summary"
  chart1.y_axis.title = 'Number'
  chart1.x_axis.title = ''
  data1 = Reference(ws, min_col=1, min_row=1, max_row=2, max_col=1)
  chart1.add_data(data1, titles_from_data=True)
  data2 = Reference(ws, min_col=2, min_row=1, max_row=2, max_col=2)
  chart1.add_data(data2, titles_from_data=True)
  ws.add_chart(chart1, CHART_LOCATION)

class Summary:
  def __init__(self, pass_count, fail_count, result_count, length_array):
    self.pass_count = pass_count
    self.fail_count = fail_count
    self.result_count = result_count
    self.length_array = length_array

def add_trend_line():
  line_props = LineProperties(solidFill = '1890ff', prstDash = 'dash', w = 15010)
  g_props = GraphicalProperties(ln=line_props)
  linear_trendline = Trendline(spPr=g_props, forward = 1, backward = 1)
  return linear_trendline

def set_data_line(line):
  line.marker.symbol = "circle"
  line.marker.size = 6
  line.marker.graphicalProperties.solidFill = "e6fffb" # Marker filling
  line.marker.graphicalProperties.line.solidFill = "13c2c2"
  line.trendline = add_trend_line()
  line.graphicalProperties.line.noFill = True

def set_spec_line(line):
  line_props = LineProperties(solidFill = 'fa541c', w = 15010)
  g_props = GraphicalProperties(ln=line_props)
  line.trendline = Trendline(spPr=g_props, forward = 1, backward = 1)
  line.graphicalProperties.line.noFill = True
  return 0

if __name__ == "__main__":
  gen_excel()